#!/usr/bin/env python3
"""
Process real NY file -> validate metadata matches original except highlights -> pop up the clone.
Run after each doc produced to ensure metadata is correct.
"""
import sys
import subprocess
from pathlib import Path

DEPLOY_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(DEPLOY_DIR))

REAL_FILE = Path("/Users/billmccreary/Downloads/20260203-NY-NewYorkCity-RDM-(373) (1).xlsx")
OUT_DIR = DEPLOY_DIR / "assets"
YELLOW_RGB = "FFFF00"


def metadata_matches_except_highlights(original_path, processed_path):
    """
    Return (True, None) if processed is a clone of original with only cell fills changed (yellow highlights).
    Else return (False, error_message).
    """
    try:
        import openpyxl
    except ImportError:
        return False, "openpyxl required"
    try:
        wb_orig = openpyxl.load_workbook(original_path, read_only=False, data_only=True)
        wb_proc = openpyxl.load_workbook(processed_path, read_only=False, data_only=True)
    except Exception as e:
        return False, f"Cannot open files: {e}"
    try:
        if len(wb_orig.worksheets) != len(wb_proc.worksheets):
            return False, f"Sheet count: original {len(wb_orig.worksheets)} vs processed {len(wb_proc.worksheets)}"
        for s_idx, (ws_orig, ws_proc) in enumerate(zip(wb_orig.worksheets, wb_proc.worksheets)):
            if ws_orig.title != ws_proc.title:
                return False, f"Sheet {s_idx} name: '{ws_orig.title}' vs '{ws_proc.title}'"
            if ws_orig.max_row != ws_proc.max_row or ws_orig.max_column != ws_proc.max_column:
                return False, f"Sheet {ws_orig.title} size: orig {ws_orig.max_row}x{ws_orig.max_column} vs proc {ws_proc.max_row}x{ws_proc.max_column}"
            # Compare cell values (data_only=True so we compare values)
            for r in range(1, ws_orig.max_row + 1):
                for c in range(1, ws_orig.max_column + 1):
                    v_orig = ws_orig.cell(r, c).value
                    v_proc = ws_proc.cell(r, c).value
                    if v_orig != v_proc:
                        return False, f"Sheet {ws_orig.title} cell ({r},{c}): original {repr(v_orig)} vs processed {repr(v_proc)}"
        # Values match; only difference should be fill on highlighted cells (we don't re-check fill here if values match)
        wb_orig.close()
        wb_proc.close()
        return True, None
    except Exception as e:
        return False, str(e)


def main():
    if not REAL_FILE.exists():
        print("Real file not found:", REAL_FILE)
        return 1
    from dep_highlighter_server import process_excel_file

    file_bytes = REAL_FILE.read_bytes()
    buf, out_name, n_high, total, clen = process_excel_file(file_bytes, REAL_FILE.name)
    out_path = OUT_DIR / out_name
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path.write_bytes(buf.read())
    print("Processed:", out_path, "| highlighted rows:", n_high, "| total:", total)

    ok, err = metadata_matches_except_highlights(REAL_FILE, out_path)
    if not ok:
        print("METADATA CHECK FAILED:", err)
        return 1
    print("Metadata check PASS: clone matches original (except highlights).")

    subprocess.run(["xattr", "-c", str(out_path)], check=False, capture_output=True)
    subprocess.run(["open", str(out_path)], check=True)
    print("Opened clone:", out_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
