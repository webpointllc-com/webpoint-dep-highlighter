#!/usr/bin/env python3
"""
DEP Highlighter Backend Server
Same logic as working Windows app: Parcel Notes + Parcel Number, consecutive DEP rows highlighted.
All processing runs on the server; the browser only uploads the file and downloads the result.
User device (Mac, Windows, etc.) does not affect processing — only the server does the work.
Deploy: push -> GitHub Action -> Render deploy hook -> live.
"""

from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import io
import os
import sys
import platform
import tempfile
from pathlib import Path
import traceback
from datetime import datetime

# Max upload size (50MB) - helps avoid Render memory/timeout issues
MAX_FILE_SIZE = 50 * 1024 * 1024

app = Flask(__name__, static_folder=None)

CORS_ORIGINS = [
    "https://webpointllc.com",
    "https://www.webpointllc.com",
    "http://localhost:5000",
    "http://localhost:5001",
    "http://127.0.0.1:5000",
    "http://127.0.0.1:5001",
]
CORS(app, origins=CORS_ORIGINS, supports_credentials=False)

_BASE = Path(__file__).resolve().parent
_FRONTEND_HTML = _BASE / "static" / "dep_highlighter.html"
if not _FRONTEND_HTML.exists():
    _FRONTEND_HTML = _BASE.parent / "frontend" / "dep_highlighter.html"

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Fixed columns per spec: Column D = Parcel Number, Column H = Parcel Notes (no auto-detect)
COL_D_PARCEL = 3   # 0-based index (Excel column D = 4th column)
COL_H_NOTES = 7    # 0-based index (Excel column H = 8th column)
MIN_COLUMNS = 8    # Need at least D and H


def _parcel_val(val):
    """Normalize parcel value for comparison."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    s = str(val).strip()
    return "" if s.upper() == "NAN" else s


def _notes_has_dep(val):
    """True if cell value contains the note 'DEP' (Parcel Notes column H)."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return False
    return "DEP" in str(val).strip().upper()


def highlight_logic(df):
    """
    Core logic (exact spec):
    - Column D = Parcel Number (same value).
    - Column H = Parcel Notes (must contain the note "DEP").
    Highlight a row ONLY when BOTH are true:
    (1) Two or more consecutive rows have the same value in column D, and
    (2) Column H has the note DEP on those rows.
    No values added; only highlight existing rows that meet both rules.
    """
    df = df.copy()
    if df.shape[1] < MIN_COLUMNS:
        raise ValueError(
            f"Sheet must have at least 8 columns (Column D = Parcel Number, Column H = Parcel Notes). Found {df.shape[1]} columns."
        )
    parcel = df.iloc[:, COL_D_PARCEL].fillna("").astype(str).str.strip()
    notes = df.iloc[:, COL_H_NOTES].fillna("").astype(str).str.strip().str.upper()
    dep_mask = notes.str.contains("DEP", na=False)

    # Find maximal runs: consecutive rows with same column D value and column H contains DEP (length >= 2)
    runs = []
    i = 0
    while i < len(df):
        if not dep_mask.iloc[i]:
            i += 1
            continue
        p = _parcel_val(parcel.iloc[i])
        if not p:
            i += 1
            continue
        j = i
        while j < len(df) and dep_mask.iloc[j] and _parcel_val(parcel.iloc[j]) == p:
            j += 1
        if j - i >= 2:
            runs.append((i, j - 1))
        i = j

    flags = [False] * len(df)
    for start, end in runs:
        for idx in range(start, end + 1):
            flags[idx] = True
    df["_highlight"] = flags
    return df


def process_excel_file(file_bytes, original_filename):
    """Read with pandas, run highlight logic, build a NEW workbook from scratch with yellow fill.
    Output is minimal valid xlsx so Excel on Mac and Windows opens it.
    Supports NY RDM-style docs: header on row 6, Tax ID / Bill ID columns."""
    file_ext = Path(original_filename).suffix.lower()
    if file_ext == ".xls":
        raise ValueError("Old .xls is not supported. Save as .xlsx or .xlsm.")

    buf = io.BytesIO(file_bytes)
    df = None
    header_row_used = 0
    for header_row in (0, 5, 4, 6, 3, 7):
        try:
            buf.seek(0)
            df_try = pd.read_excel(buf, engine="openpyxl", sheet_name=0, header=header_row)
            if df_try is None or len(df_try) == 0:
                continue
            if df_try.shape[1] >= MIN_COLUMNS:
                df = df_try
                header_row_used = header_row
                break
        except Exception:
            continue
    if df is None:
        try:
            buf.seek(0)
            df = pd.read_excel(buf, engine="openpyxl", sheet_name=0)
        except Exception as e:
            raise ValueError(f"Cannot read Excel file. Is it a valid .xlsx or .xlsm? Details: {e}")
        if df is None or len(df) == 0:
            raise ValueError("The file has no data rows.")
        if df.shape[1] < MIN_COLUMNS:
            raise ValueError(
                "Sheet must have at least 8 columns. Column D = Parcel Number, Column H = Parcel Notes."
            )

    df_processed = highlight_logic(df)
    rows_to_highlight = set()
    for pandas_idx, row in df_processed.iterrows():
        if row["_highlight"]:
            rows_to_highlight.add(int(pandas_idx))

    base_name = Path(original_filename).stem
    extension = Path(original_filename).suffix
    output_filename = f"{base_name}_Highlighted{extension}"

    # CLONE ONLY: Load original workbook, apply yellow fill ONLY to highlighted rows.
    # No rebuild. 100% preservation of VBA macros, hidden/visible code, all sheets, metadata.
    # All submitted files have exactly Column D = Parcel Number, Column H = Parcel Notes.
    try:
        buf.seek(0)
        wb = openpyxl.load_workbook(buf, keep_vba=(file_ext == ".xlsm"), data_only=False)
        sheet = wb.worksheets[0]
        # pandas index i -> Excel row (1-based): header at row header_row_used+1, data starts header_row_used+2
        for pandas_idx in rows_to_highlight:
            excel_row = header_row_used + 2 + int(pandas_idx)
            if 1 <= excel_row <= sheet.max_row:
                for col in range(1, sheet.max_column + 1):
                    sheet.cell(excel_row, col).fill = YELLOW_FILL
        tmp = None
        try:
            fd, tmp = tempfile.mkstemp(suffix=extension)
            os.close(fd)
            wb.save(tmp)
            with open(tmp, "rb") as f:
                output_bytes = f.read()
        finally:
            if tmp and os.path.exists(tmp):
                try:
                    os.unlink(tmp)
                except OSError:
                    pass
    except Exception as e:
        raise ValueError(
            f"Cannot preserve workbook (VBA/macros/hidden code must be kept). Clone failed: {e}. "
            "File must be valid .xlsx or .xlsm with at least 8 columns (Column D = Parcel Number, Column H = Parcel Notes)."
        ) from e

    return io.BytesIO(output_bytes), output_filename, len(rows_to_highlight), len(df), len(output_bytes)


@app.errorhandler(500)
def handle_500(e):
    msg = str(e) if str(e) else "Internal server error"
    return jsonify({"error": msg, "details": msg}), 500


@app.errorhandler(Exception)
def handle_exception(e):
    msg = str(e) or "Unexpected server error"
    return jsonify({"error": msg, "details": msg}), 500


@app.route("/health", methods=["GET"])
def health_check():
    return jsonify({
        "status": "ok",
        "service": "DEP Highlighter",
        "timestamp": datetime.now().isoformat(),
        "python_version": platform.python_version(),
        "openpyxl_installed": "openpyxl" in sys.modules,
        "temp_writable": os.access("/tmp", os.W_OK),
        "cwd": os.getcwd(),
    })


@app.route("/process", methods=["POST"])
def process_file():
    temp_path = None
    try:
        print(f"[{datetime.now().isoformat()}] === Processing started ===", flush=True)

        if "file" not in request.files:
            print("400: No file in request.files", flush=True)
            return jsonify({"error": "No file provided", "details": "No file provided. The upload form did not include a file. Try selecting a file again and click Process."}), 400
        file = request.files["file"]
        if file.filename == "" or not (file.filename or "").strip():
            print("400: Empty filename", flush=True)
            return jsonify({"error": "No file selected", "details": "No file selected. Please choose an Excel file (.xlsx or .xlsm) and try again."}), 400

        file_ext = Path(file.filename).suffix.lower()
        if file_ext not in (".xlsx", ".xlsm", ".xls"):
            print(f"400: Invalid file type: {file_ext}", flush=True)
            return jsonify({"error": "Invalid file type. Use .xlsx or .xlsm", "details": "Invalid file type. Use .xlsx or .xlsm"}), 400
        if file_ext == ".xls":
            print("400: .xls not supported", flush=True)
            return jsonify({"error": "Old .xls not supported. Save as .xlsx or .xlsm", "details": "Old .xls not supported. Save as .xlsx or .xlsm"}), 400

        file_bytes = file.read()
        file_size = len(file_bytes)
        print(f"File received: {file.filename}, size: {file_size} bytes ({file_size / 1024 / 1024:.2f} MB)", flush=True)

        if not file_bytes or file_size < 100:
            print(f"400: File too small: {file_size} bytes", flush=True)
            return jsonify({"error": "File is empty or too small. Use a valid .xlsx or .xlsm file.", "details": "File is empty or too small (under 100 bytes). Use a valid .xlsx or .xlsm file."}), 400

        if file_size > MAX_FILE_SIZE:
            msg = f"File too large: {file_size / 1024 / 1024:.1f} MB. Max {MAX_FILE_SIZE // (1024*1024)} MB."
            print(msg, file=sys.stderr, flush=True)
            return jsonify({"error": msg, "details": msg}), 413

        # Optional temp save for Render diagnostics (per doc: "File system issues - temp file writes failing")
        try:
            safe_name = "".join(c if c.isalnum() or c in "._-" else "_" for c in file.filename)
            temp_path = os.path.join(tempfile.gettempdir(), safe_name)
            print(f"Saving to temp file: {temp_path}", flush=True)
            with open(temp_path, "wb") as f:
                f.write(file_bytes)
            print(f"Saved to: {temp_path}", flush=True)
        except Exception as te:
            print(f"Temp save failed (continuing in-memory): {te}", file=sys.stderr, flush=True)

        print("Loading workbook with openpyxl...", flush=True)
        output_buffer, output_filename, highlighted_count, total_rows, content_length = process_excel_file(
            file_bytes, file.filename
        )
        print(f"Processing complete. Rows: {total_rows}, highlighted: {highlighted_count}. Sending file: {output_filename}", flush=True)

        return send_file(
            output_buffer,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=output_filename,
        )
    except ValueError as ve:
        print(f"ValueError: {ve}", file=sys.stderr, flush=True)
        return jsonify({"error": str(ve), "details": str(ve)}), 400
    except MemoryError as e:
        error_msg = f"MEMORY ERROR: {e}"
        print(error_msg, file=sys.stderr, flush=True)
        return jsonify({"error": "File too large for processing", "details": error_msg}), 413
    except Exception as e:
        msg = str(e).strip() or "Processing failed"
        if len(msg) > 400:
            msg = msg[:397] + "..."
        error_msg = f"ERROR: {e}\n{traceback.format_exc()}"
        print(error_msg, file=sys.stderr, flush=True)
        return jsonify({"error": msg, "details": msg}), 500
    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
                print(f"Cleaned up {temp_path}", flush=True)
            except Exception:
                pass


@app.route("/", methods=["GET"])
def index():
    if _FRONTEND_HTML.exists():
        return send_file(str(_FRONTEND_HTML), mimetype="text/html")
    return jsonify({
        "service": "Webpoint LLC - DEP Highlighter API",
        "version": "1.0.0",
        "endpoints": {"/health": "GET", "/process": "POST"},
    })


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("DEBUG", "false").lower() == "true"
    app.run(host="0.0.0.0", port=port, debug=debug)
